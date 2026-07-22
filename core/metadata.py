import os
from pathlib import Path
from datetime import datetime


async def extract_image_metadata(filepath: str) -> dict:
    filepath = Path(filepath)
    result = {
        "file": str(filepath),
        "type": "image",
        "size_bytes": filepath.stat().st_size if filepath.exists() else 0,
    }

    if not filepath.exists():
        result["error"] = "File not found"
        return result

    try:
        from PIL import Image
        from PIL.ExifTags import TAGS, GPSTAGS

        img = Image.open(filepath)
        result["format"] = img.format
        result["mode"] = img.mode
        result["width"] = img.width
        result["height"] = img.height

        exif_data = img._getexif()
        if exif_data:
            exif = {}
            for tag_id, value in exif_data.items():
                tag_name = TAGS.get(tag_id, tag_id)
                if tag_name == "GPSInfo":
                    gps_data = {}
                    for gps_tag_id in value:
                        gps_tag_name = GPSTAGS.get(gps_tag_id, gps_tag_id)
                        gps_data[gps_tag_name] = str(value[gps_tag_id])
                    exif[tag_name] = gps_data
                elif isinstance(value, bytes):
                    exif[tag_name] = f"<{len(value)} bytes>"
                else:
                    exif[tag_name] = str(value)
            result["exif"] = exif

            if "GPSInfo" in exif:
                gps = _parse_gps(exif["GPSInfo"])
                if gps:
                    result["gps_parsed"] = gps
                    result["maps_link"] = f"https://www.google.com/maps?q={gps['lat']},{gps['lon']}"

        img.close()
    except ImportError:
        result["error"] = "Pillow not installed"
    except Exception as e:
        result["error"] = str(e)

    return result


def _parse_gps(gps_data: dict) -> dict | None:
    try:
        def _convert(coord_str, ref_str):
            parts = coord_str.strip("()").split(",")[:3]
            deg = float(parts[0].split("=")[-1].strip())
            min_ = float(parts[1].split("=")[-1].strip())
            sec = float(parts[2].split("=")[-1].strip().rstrip(")"))
            val = deg + min_ / 60 + sec / 3600
            if ref_str and ("S" in ref_str.upper() or "W" in ref_str.upper()):
                val *= -1
            return val

        if "GPSLatitude" not in gps_data or "GPSLatitudeRef" not in gps_data:
            return None

        lat = _convert(gps_data["GPSLatitude"], str(gps_data.get("GPSLatitudeRef", "N")))
        lon = _convert(gps_data["GPSLongitude"], str(gps_data.get("GPSLongitudeRef", "E")))

        if lat and lon:
            return {"lat": round(lat, 6), "lon": round(lon, 6)}
    except Exception:
        pass
    return None


async def extract_pdf_metadata(filepath: str) -> dict:
    filepath = Path(filepath)
    result = {
        "file": str(filepath),
        "type": "pdf",
        "size_bytes": filepath.stat().st_size if filepath.exists() else 0,
    }

    if not filepath.exists():
        result["error"] = "File not found"
        return result

    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(filepath)
        result["pages"] = len(reader.pages)
        result["encrypted"] = reader.is_encrypted
        result["metadata"] = {}

        if reader.metadata:
            for key in reader.metadata:
                cleaned_key = key.lstrip("/")
                result["metadata"][cleaned_key] = str(reader.metadata[key])
    except ImportError:
        result["error"] = "PyPDF2 not installed"
    except Exception as e:
        result["error"] = str(e)

    return result


async def extract_docx_metadata(filepath: str) -> dict:
    filepath = Path(filepath)
    result = {
        "file": str(filepath),
        "type": "docx",
        "size_bytes": filepath.stat().st_size if filepath.exists() else 0,
    }

    if not filepath.exists():
        result["error"] = "File not found"
        return result

    try:
        from docx import Document
        doc = Document(filepath)
        props = doc.core_properties
        result["metadata"] = {
            "author": props.author,
            "category": props.category,
            "comments": props.comments,
            "content_status": props.content_status,
            "created": str(props.created) if props.created else None,
            "identifier": props.identifier,
            "keywords": props.keywords,
            "language": props.language,
            "last_modified_by": props.last_modified_by,
            "last_printed": str(props.last_printed) if props.last_printed else None,
            "modified": str(props.modified) if props.modified else None,
            "revision": props.revision,
            "subject": props.subject,
            "title": props.title,
            "version": props.version,
        }
    except ImportError:
        result["error"] = "python-docx not installed"
    except Exception as e:
        result["error"] = str(e)

    return result


async def extract_xlsx_metadata(filepath: str) -> dict:
    filepath = Path(filepath)
    result = {
        "file": str(filepath),
        "type": "xlsx",
        "size_bytes": filepath.stat().st_size if filepath.exists() else 0,
    }

    if not filepath.exists():
        result["error"] = "File not found"
        return result

    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True)
        result["sheets"] = wb.sheetnames
        result["metadata"] = {
            "creator": wb.properties.creator,
            "title": wb.properties.title,
            "description": wb.properties.description,
            "subject": wb.properties.subject,
            "category": wb.properties.category,
            "created": str(wb.properties.created) if wb.properties.created else None,
            "modified": str(wb.properties.modified) if wb.properties.modified else None,
            "last_modified_by": wb.properties.last_modified_by or wb.properties.lastModifiedBy,
            "keywords": wb.properties.keywords,
        }
        wb.close()
    except ImportError:
        result["error"] = "openpyxl not installed"
    except Exception as e:
        result["error"] = str(e)

    return result


async def extract_metadata(filepath: str) -> dict:
    ext = Path(filepath).suffix.lower()
    if ext in (".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff", ".webp"):
        return await extract_image_metadata(filepath)
    elif ext == ".pdf":
        return await extract_pdf_metadata(filepath)
    elif ext == ".docx":
        return await extract_docx_metadata(filepath)
    elif ext == ".xlsx":
        return await extract_xlsx_metadata(filepath)
    else:
        return {"file": filepath, "error": f"Unsupported file type: {ext}"}
